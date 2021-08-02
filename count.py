print('Initializing...')

import pandas as pd
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, colors, Color, Alignment, PatternFill, GradientFill, Border, Side
from openpyxl.styles import NamedStyle
import matplotlib.pyplot as plt

wb = Workbook('Reporte.xlsx')
wb.save('Reporte.xlsx')

list_of_boxes = list(['box_1291.xlsx', 'box_1292.xlsx', 'box_1293.xlsx', 'box_1294.xlsx', 'box_1295.xlsx', 'box_1296.xlsx', 'box_1297.xlsx', 'box_1298.xlsx', 'box_1299.xlsx', 'box_1300.xlsx'])                                                                    # modify with range of boxes
print('Boxes processed:')    

for file_name in list_of_boxes: 
        
    df_all = pd.concat(pd.read_excel(file_name, sheet_name=None, usecols="A:C"))
    #print(df_all)

    #df2 = df_all.groupby(['name','direction']).sum('quantity')                                                      # crear una versión para contar el número de respuestas
    df2 = df_all.groupby(['name']).sum('quantity')
    #print(df2)
    #df3 = df2.sort_values(['name','quantity'], ascending=[True, False])
    df3 = df2.sort_values(['quantity'], ascending=False)
    #print(df3)

    wb = load_workbook('Reporte.xlsx')    

    writer = pd.ExcelWriter('Reporte.xlsx', engine='openpyxl', mode='a')
    df3.to_excel(writer, sheet_name=file_name)
    writer.save()

    print(file_name)

df_total = pd.concat(pd.read_excel('Reporte.xlsx', sheet_name=[1, 2, 3, 4, 5, 6, 7, 8, 9, 10]))                 # change this range when adding new boxes
df_tot1 = df_total.groupby(['name']).sum('quantity')
df_tot2 = df_tot1.sort_values(['quantity'], ascending=False)

wb = load_workbook('Reporte.xlsx')    

writer = pd.ExcelWriter('Reporte.xlsx', engine='openpyxl', mode='a')
df_tot2.to_excel(writer, 'Total')
writer.save()

wb = load_workbook('Reporte.xlsx')
del wb['Sheet']
wb.save('Reporte.xlsx')

for sheet in wb.worksheets:
    
    sheet['D1'] = 'Total Names:'
    sheet['F1'] = 'Total Letters:'
    sheet['E1'] = '=COUNTA(A:A) - 1'
    sheet['G1'] = '=SUM(B:B)'

    aut = sheet['D1']
    aut.font = Font(color='00FF6600', size=12, bold=True)
    aut.fill = PatternFill(bgColor='00000000', fill_type = "solid")
    sheet.column_dimensions['D'].width = 20
    let = sheet['F1']
    let.font = Font(color='00FF6600', size=12, bold=True)
    let.fill = PatternFill(bgColor='00000000', fill_type = "solid")
    sheet.column_dimensions['F'].width = 20

    authors_cell = sheet['A1']
    dir_cell = sheet['B1']
    quant_cell = sheet['C1']
    quant_col = sheet['B']
    authors_cell.font = Font(color='00FF6600', size=12, bold=True)
    dir_cell.font = Font(color='00FF6600', size=12, bold= True) 
    quant_cell.font = Font(color='00FF6600', size=12, bold= True)
    authors_cell.alignment = Alignment(horizontal='center', vertical='center')
    quant_cell.alignment = Alignment(horizontal='center', vertical='center')
    authors_cell.fill = PatternFill(bgColor='00000000', fill_type = "solid")
    dir_cell.fill = PatternFill(bgColor='00000000', fill_type = "solid")
    quant_cell.fill = PatternFill(bgColor='00000000', fill_type = "solid")
    sheet.column_dimensions['A'].width = 50
    sheet.column_dimensions['B'].width = 20


    thin_border = Border(left=Side(style='thin'), 
                        right=Side(style='thin'), 
                        top=Side(style='thin'), 
                        bottom=Side(style='thin'))
    quant_col = sheet['B']

    for row in quant_col:
        row.alignment = Alignment(horizontal='center')
        row.border = thin_border 

    for row in sheet.iter_rows():
        for cell in row:
            if cell.value == ('sin_identificar'):
                cell.fill = PatternFill(start_color='0000FF00', end_color='0000FF00', fill_type = "solid") 

    #for row in sheet.iter_rows():
        #for cell in row:
            #if value in cell.value = 1:
                #cell.fill = PatternFill(start_color='0000FF00', end_color='0000FF00', fill_type = "solid")

    sheet.freeze_panes = 'B2'
    wb.save('Reporte.xlsx')

wb.save('Reporte.xlsx')

print('Creating graph...')

df_final = pd.read_excel('Reporte.xlsx', sheet_name='Total', usecols='A:C')
limit = 15
df_f = df_final[:limit]
#print(df_f)

df_f.plot.bar(x='name', color='orangered')
plt.xticks(rotation=25, fontsize=5)
plt.ylabel('Cartas')
#plt.show()
plt.savefig('autores.png', dpi=150)

wb = openpyxl.load_workbook('Reporte.xlsx')
wc = wb.create_sheet('Gráfica')

img = openpyxl.drawing.image.Image('autores.png')
img.anchor = 'A1'
wc.add_image(img) 

wb.save('Reporte.xlsx')

#wb.open('Reporte.xlsx')

print('Complete!')
